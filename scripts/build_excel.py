# -*- coding: utf-8 -*-
import json, re, difflib
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX = r'C:\Users\User\Downloads\PRDB PERKAPITA 2025 (1).xlsx'
DATA = r'C:\Users\User\AppData\Local\Temp\claude\D--RAFTING-WEB\78ab370f-99cb-4904-9184-71751db79aed\scratchpad\pdrb_data.json'

with open(DATA, encoding='utf-8') as fh:
    tables = json.load(fh)

def norm(s):
    s = s.lower()
    s = s.replace('kabupaten', 'kab')
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

# lookup: normalized name -> list of (provinsi, name, vals)
lookup = {}
for t in tables:
    for name, vals, is_sum in t['rows']:
        if is_sum:
            continue
        lookup.setdefault(norm(name), []).append((t['provinsi'], name, vals))

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

matched, unmatched, fuzzy_used = 0, [], []
keys = list(lookup.keys())
for r in range(2, ws.max_row + 1):
    excel_name = ws.cell(r, 1).value
    if not excel_name:
        continue
    k = norm(str(excel_name))
    cands = lookup.get(k)
    if not cands:
        close = difflib.get_close_matches(k, keys, n=2, cutoff=0.85)
        if len(close) >= 1:
            cands = lookup[close[0]]
            fuzzy_used.append((excel_name, cands[0][1]))
    if not cands:
        unmatched.append(excel_name)
        continue
    # prefer entry with a 2025 value
    cands_with = [c for c in cands if c[2][4] is not None]
    pick = cands_with[0] if cands_with else cands[0]
    if len(cands_with) > 1:
        print('DUPLICATE with values:', excel_name, [(c[0], c[2][4]) for c in cands_with])
    v2025 = pick[2][4]
    cell = ws.cell(r, 2)
    cell.value = v2025
    cell.number_format = '#,##0'
    matched += 1

print('matched:', matched, '| unmatched:', len(unmatched))
for u in unmatched:
    print('  UNMATCHED:', u)
for a, b in fuzzy_used:
    print('  fuzzy:', a, '<-', b)

# ---- full data sheet ----
title = 'Data PDF 2021-2025'
if title in wb.sheetnames:
    del wb[title]
ws2 = wb.create_sheet(title)

ARIAL = 'Arial'
hdr_font = Font(name=ARIAL, size=10, bold=True)
body_font = Font(name=ARIAL, size=10)
sum_font = Font(name=ARIAL, size=10, bold=True, italic=True)
hdr_fill = PatternFill('solid', fgColor='D9E1F2')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws2['A1'] = 'PDRB Per Kapita Atas Dasar Harga Berlaku Menurut Kabupaten/Kota (ribu rupiah), 2021-2025'
ws2['A1'].font = Font(name=ARIAL, size=11, bold=True)
ws2['A2'] = 'Sumber: BPS, Produk Domestik Regional Bruto Kabupaten/Kota di Indonesia 2021-2025, Tabel 153-190. * angka sementara, ** angka sangat sementara, "-" data belum tersedia'
ws2['A2'].font = Font(name=ARIAL, size=9, italic=True)

headers = ['Tabel', 'Provinsi', 'Kabupaten/Kota', '2021', '2022', '2023', '2024*', '2025**']
row = 4
for c, h in enumerate(headers, 1):
    cell = ws2.cell(row, c, h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

row += 1
for t in tables:
    for name, vals, is_sum in t['rows']:
        ws2.cell(row, 1, t['tabel'])
        ws2.cell(row, 2, t['provinsi'])
        ws2.cell(row, 3, name)
        for i, v in enumerate(vals):
            cell = ws2.cell(row, 4 + i)
            if v is None:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = v
                cell.number_format = '#,##0'
        f = sum_font if is_sum else body_font
        for c in range(1, 9):
            ws2.cell(row, c).font = f
            ws2.cell(row, c).border = border
        row += 1

widths = [7, 22, 34, 11, 11, 11, 11, 11]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = 'A5'

wb.save(XLSX)
print('saved:', XLSX)
